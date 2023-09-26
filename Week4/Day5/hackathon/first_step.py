#to get more info on openpyx1
# https://www.datacamp.com/tutorial/python-excel-tutorial


import openpyxl 
import os 
import pandas as pd

dir_path = os.path.dirname(os.path.realpath(__file__))

# access to data:
wb = openpyxl.load_workbook(dir_path + r"\\energy_data.xlsx", "r")

ws_hydro = wb['Hydro Generation - TWh']
ws_ren = wb['Renewable power - TWh']
ws_nuc = wb['Nuclear Generation - TWh']

ws_oil = wb['Elec Gen from Oil']
ws_gas = wb['Elec Gen from Gas']
ws_coal = wb['Elec Gen from Coal']
ws_other = wb['Elec Gen from Other'] 


class EnergyData:
    def __init__(self, country):
        self.country = country

    def fetch_data65(self, worksheet, start_year, end_year):
        values65 = []

        start_col = start_year - 1965 + 2
        end_col = end_year - 1965 + 2
    
        for row in worksheet.iter_rows(min_row=3, max_row=109, min_col=start_col, max_col=end_col):
            cell_in_column_A = worksheet.cell(row=row[0].row, column=1) 
        if cell_in_column_A.value == self.country:
            row_values = [cell.value for cell in row]
            values65.append(row_values)
        
        return values65
    
    def fetch_data85(self,  worksheet, start_year, end_year):
        values85 = []

        start_col = start_year - 1985 + 2
        end_col = end_year - 1985 + 2
    
        for row in worksheet.iter_rows(min_row=3, max_row=109, min_col=start_col, max_col=end_col):
            cell_in_column_A = worksheet.cell(row=row[0].row, column=1) 
        if cell_in_column_A.value == self.country:
            row_values = [cell.value for cell in row]
            values85.append(row_values)
        
        return values85

    def histdata(self, start_year, end_year):
        hy_values = self.fetch_data65(ws_hydro, start_year, end_year)
        ren_values = self.fetch_data65(ws_ren, start_year, end_year)
        nuc_values = self.fetch_data65(ws_nuc, start_year, end_year)
        oil_values = self.fetch_data85(ws_oil, start_year, end_year)
        gas_values = self.fetch_data85(ws_gas, start_year, end_year)
        coal_values = self.fetch_data85(ws_coal, start_year, end_year)
        other_values = self.fetch_data85(ws_other, start_year, end_year)

        if not coal_values and not gas_values:
            print(f"No values found for {self.country}.")
            return None
    
        date = list(range(start_year, end_year+1))
        fuelmix = ["Hydro", "Renewables", "Nuclear", "Oil", "Gas", "Coal", "Other"]

        table = [date]
        for i, mix in enumerate(fuelmix):
            row_values = [fuel]
            row_values.extend(hy_values[i])
            row_values.extend(ren_values[i])
            row_values.extend(nuc_values[i])
            row_values.extend(oil_values[i])
            row_values.extend(gas_values[i])
            row_values.extend(coal_values[i])
            row_values.extend(other_values[i])
            table.append(row_values)
    
        df = pd.DataFrame(table[1:], columns = table[0])
        return df
    

a = EnergyData('Brazil')
print(a.histdata(2018, 2022))


# return date, hy_row_values, ren_row_values, nuc_row_values, oil_row_values, gas_row_values, coal_row_values, other_row_values 


     # def histdata(self, start_year, end_year):
    #     hy_values = []
    #     ren_values = []
    #     nuc_values = []

    #     start_col = start_year - 1965 + 2
    #     end_col = end_year - 1965 + 2
    #     # hydro
    #     for row in ws_hydro.iter_rows(min_row=3, max_row=109, min_col=start_col, max_col=end_col):
    #         cell_in_column_A = ws_hydro.cell(row=row[0].row, column=1) 
    #         if cell_in_column_A.value == self.country:
    #             hy_row_values = [cell.value for cell in row]
    #             hy_values.append(hy_row_values)
    #     # renewable
    #     for row in ws_ren.iter_rows(min_row=3, max_row=109, min_col=start_col, max_col=end_col):
    #         cell_in_column_A = ws_ren.cell(row=row[0].row, column=1)  
    #         if cell_in_column_A.value == self.country:
    #             ren_row_values = [cell.value for cell in row]
    #             ren_values.append(ren_row_values)
    #     # nuclear
    #     for row in ws_nuc.iter_rows(min_row=3, max_row=109, min_col=start_col, max_col=end_col):
    #         cell_in_column_A = ws_nuc.cell(row=row[0].row, column=1)  
    #         if cell_in_column_A.value == self.country:
    #             nuc_row_values = [cell.value for cell in row]
    #             nuc_values.append(nuc_row_values)
    #     #data for other sources starts from 1985, so the baseyear changes:
    #     oil_values = []
    #     gas_values = []
    #     coal_values = []
    #     other_values = []

    #     start_col2 = start_year - 1985 + 2
    #     end_col2 = end_year - 1985 + 2   

    #     #oil   
    #     for row in ws_oil.iter_rows(min_row=3, max_row=109, min_col=start_col2, max_col=end_col2):
    #         cell_in_column_A = ws_oil.cell(row=row[0].row, column=1)  
    #         if cell_in_column_A.value == self.country:
    #             oil_row_values = [cell.value for cell in row]
    #             oil_values.append(oil_row_values)
    #     #gas   
    #     for row in ws_gas.iter_rows(min_row=3, max_row=109, min_col=start_col2, max_col=end_col2):
    #         cell_in_column_A = ws_gas.cell(row=row[0].row, column=1)  
    #         if cell_in_column_A.value == self.country:
    #             gas_row_values = [cell.value for cell in row]
    #             gas_values.append(gas_row_values)

    #     #coal   
    #     for row in ws_coal.iter_rows(min_row=3, max_row=109, min_col=start_col2, max_col=end_col2):
    #         cell_in_column_A = ws_coal.cell(row=row[0].row, column=1)  
    #         if cell_in_column_A.value == self.country:
    #             coal_row_values = [cell.value for cell in row]
    #             coal_values.append(coal_row_values)
        
    #     #other   
    #     for row in ws_other.iter_rows(min_row=3, max_row=109, min_col=start_col2, max_col=end_col2):
    #         cell_in_column_A = ws_other.cell(row=row[0].row, column=1)  
    #         if cell_in_column_A.value == self.country:
    #             other_row_values = [cell.value for cell in row]
    #             other_values.append(other_row_values)
        
    #     #forming a table
    #     date = []
    #     for i in range(start_year, end_year+1): 
    #         date.append(i)
        
    #     fuelmix = ["Hydro", "Renewables", "Nuclear", "Oil", "Gas", "Coal", "Other"]

    #     table = []
    #     table.append(date)
    #     for i in range(len(fuelmix)+1):
    #         table.append([fuelmix[i], hy_row_values[i], ren_row_values[i], nuc_row_values[i], oil_row_values[i], gas_row_values[i], coal_row_values[i], other_row_values[i]])

    #     df = pd.DataFrame(table[1:], columns = table[0])
    #     if not coal_values and not gas_values:
    #         print(f"No values found for {self.country}.")
    #     else:
    #         return df