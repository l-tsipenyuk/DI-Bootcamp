import openpyxl 
import os 
import pandas as pd
import psycopg2
import psycopg2.extras

# access to data in Excel:
dir_path = os.path.dirname(os.path.realpath(__file__))
wb = openpyxl.load_workbook(dir_path + r"\\energy_data.xlsx", "r")

ws_hydro = wb['Hydro Generation - TWh']
ws_ren = wb['Renewable power - TWh']
ws_nuc = wb['Nuclear Generation - TWh']

ws_oil = wb['Elec Gen from Oil']
ws_gas = wb['Elec Gen from Gas']
ws_coal = wb['Elec Gen from Coal']
ws_other = wb['Elec Gen from Other'] 

ws_total = wb['Electricity Generation']

class EnergyData:
    def __init__(self, country):
        self.country = country

    def valid_country(self):
        for row in range(5, 110): 
            cell_in_column_A = ws_hydro.cell(row=row, column=1).value
            if cell_in_column_A == self.country:
                return True 
        return False  

    #methods to collect the generation mix from various spreadsheets

    def fetch_data65(self, worksheet, start_year, end_year):
        values65 = []

        start_col = start_year - 1965 + 2
        end_col = end_year - 1965 + 2

        for row in worksheet.iter_rows(min_row=3, max_row=109, min_col=start_col, max_col=end_col):
            cell_in_column_A = worksheet.cell(row=row[0].row, column=1)
            if cell_in_column_A.value == self.country:
                row_values = [cell.value for cell in row]
                values65.append(row_values)

        return values65

    def fetch_data85(self,  worksheet, start_year, end_year):
        values85 = []

        start_col = start_year - 1985 + 2
        end_col = end_year - 1985 + 2

        for row in worksheet.iter_rows(min_row=3, max_row=109, min_col=start_col, max_col=end_col):
            cell_in_column_A = worksheet.cell(row=row[0].row, column=1)
            if cell_in_column_A.value == self.country:
                row_values = [cell.value for cell in row]
                values85.append(row_values)

        return values85

    def histdata(self, start_year, end_year):
        hy_values = self.fetch_data65(ws_hydro, start_year, end_year)[0]
        ren_values = self.fetch_data65(ws_ren, start_year, end_year)[0]
        nuc_values = self.fetch_data65(ws_nuc, start_year, end_year)[0]
        oil_values = self.fetch_data85(ws_oil, start_year, end_year)[0]
        gas_values = self.fetch_data85(ws_gas, start_year, end_year)[0]
        coal_values = self.fetch_data85(ws_coal, start_year, end_year)[0]
        other_values = self.fetch_data85(ws_other, start_year, end_year)[0]
        
        date = list(range(start_year, end_year+1))
        fuelmix = ["Hydro", "Renewables", "Nuclear", "Oil", "Gas", "Coal", "Other"]


        table = [[" "] + fuelmix]
        for i, year in enumerate(date):
            row_values = [year] + [hy_values[i], ren_values[i], nuc_values[i], oil_values[i], gas_values[i], coal_values[i], other_values[i]]
            table.append(row_values)

        df = pd.DataFrame(table[1:], columns=table[0]).set_index(' ')
        df.reset_index(drop=True)
        return df.T

# method to get the generation mix in percentage for a specific country 

    def get_the_share(self, country, year):
        table = self.histdata(year, year)
        total_sum = table[year].sum()
        table['Share, %'] = ((table[year]/ total_sum) * 100).round(1)
        table = table.sort_values(by='Share, %', ascending = False)
        return table

# methods to get the peers WORK ON THIS

    def get_the_peer(self, country, year, fuel_type):
        country_list = ["Canada", "Mexico", "US", "Argentina", "Brazil", "France", "Germany", "Italy", "Netherlands", "Poland", "Spain", "Turkey", "Ukraine", "United Kingdom", "Kazakhstan", "Russian Federation", "Iran", "Saudi Arabia", "United Arab Emirates", "Egypt", "Australia", "China", "India", "Indonesia", "Japan", "Malaysia", "South Korea", "Taiwan", "Thailand", "Vietnam"]
        
        for i in country_list:
            a = EnergyData(i)
            a.histdata(year,year)
        # table2 = self.get_the_share(country_list, year)
        # return 
 



        # total_gen_data = {}
        # fuel_type_gen_data = {}

        # year_column_total = None
        # for col_index, cell in enumerate(ws_total[3]):
        #     if cell.value == year:
        #         year_column_total = col_index + 1  
        #         break

        # year_column_hydro = None
        # for col_index, cell in enumerate(ws_hydro[3]):
        #     if cell.value == year:
        #         year_column_hydro = col_index + 1  
        #         break

        # if year_column_total is not None and year_column_hydro is not None:
        #     for row_total, row_hydro in zip(
        #         ws_total.iter_rows(min_row=4, max_row=109, min_col=year_column_total, max_col=year_column_total),
        #         ws_hydro.iter_rows(min_row=4, max_row=109, min_col=year_column_hydro, max_col=year_column_hydro)
        #     ):
        #         country_cell_total = row_total[0]
        #         country_cell_hydro = row_hydro[0]
        #         total_gen = country_cell_total.value
        #         fuel_gen = country_cell_hydro.value

        #         if total_gen is not None and fuel_gen is not None:
        #             total_gen_data[country_cell_total.value] = total_gen
        #             fuel_type_gen_data[country_cell_hydro.value] = fuel_gen

        # share_data = {}
        # for country in total_gen_data:
        #     total_gen = total_gen_data[country]
        #     fuel_type_gen = fuel_type_gen_data.get(country, 0)

        #     if total_gen is not None and total_gen != 0:
        #         share_data[country] = (fuel_type_gen / total_gen) * 100

        # country_share = share_data.get(self.country, 0)
        # sorted_countries = sorted(share_data.items(), key=lambda x: abs(x[1] - country_share))
        # result = sorted_countries[:5]

        # print(year_column_total)

        # for country, share in result:
        #     print(f"Country: {country}, Share, %: {share} ")

#methods to work with SQL Postgress

    def connect_to_postgresql(self):
        DB_NAME = "energy_mix"
        USER = "postgres"
        PASSWORD = "root"
        HOST = "localhost"
        PORT = "5432" 

        try:
            global connection
            connection = psycopg2.connect(
                dbname = DB_NAME, 
                user = USER,
                password = PASSWORD,
                host = HOST,
                port = PORT
            )
        except Exception as e:
            print(f"Error: {e}")
        
        global cursor
        cursor = connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    def create_a_table(self, table_name: str, start_year, end_year):
        try:
            self.connect_to_postgresql()
            date = list(range(start_year, end_year+1))
            query = f'''
                create table if not exists {table_name} (
                    fuel_type varchar(20) primary key,
                    {', '.join([f'"{year}" numeric' for year in date])}
                );
                '''
            cursor.execute(query) 
            connection.commit()
            return f"The table {table_name} is created."
        except Exception as e:
            None

    def add_data_to_a_table(self, table_name: str, start_year, end_year):
        try:
            self.connect_to_postgresql()
            date = list(range(start_year, end_year+1))
            data = self.histdata(start_year, end_year)
            
            all_values = []

            for fuel_type, row in data.iterrows():
                values = list(row)
                all_values.append((fuel_type, values))

            query = f'''
                insert into {table_name} (fuel_type, {', '.join([f'"{year}"' for year in date])})
                values
                {', '.join([f"('{fuel_type}', {', '.join([str(v) for v in values])})" for fuel_type, values in all_values])};
            '''

            cursor.execute(query) 
            connection.commit()
            return f"The data is added into {table_name}."
        except Exception as e:
            None
    
    def delete_the_table(self, table_name: str):
        try:
            self.connect_to_postgresql()
            query = f'''
                drop table {table_name};
            '''

            cursor.execute(query) 
            connection.commit()
            return f"The table {table_name} is deleted."
        except Exception as e:
            None


 
 

# a = EnergyData('Germany')
# print(a.get_the_share('Germany', 1995))
# print(a.get_the_peer('Germany', 2022, 'Hydro'))
# a.connect_to_postgress()
# print(a.create_a_table('energy_mix_ger', 2015, 2022))
# print(a.add_data_to_a_table('energy_mix_ger', 2015, 2022))
# print(a.delete_the_table('energy_mix_ger'))




# country_list = ["Germany", "France"]
country_list = ["Canada", "Mexico", "US", "Argentina", "Brazil", "France", "Germany", "Italy", "Netherlands", "Poland", "Spain", "Turkey", "Ukraine", "United Kingdom", "Kazakhstan", "Russian Federation", "Iran", "Saudi Arabia", "United Arab Emirates", "Egypt", "Australia", "China", "India", "Indonesia", "Japan", "Malaysia", "South Korea", "Taiwan", "Thailand", "Vietnam"]
table = {}  

for country in country_list:
    a = EnergyData(country)
    table[country] = a.get_the_share(country, 2022).rename(columns= {"Share, %": country}).drop([2022], axis=1)

result = table[country_list[0]]

for country in country_list[1:]:
    result = pd.merge(result, table[country], left_index=True, right_index=True)

result = result.reset_index()
sorted_result = result.sort_values(by= 'Hydro', axis=1)
short_sorted = sorted_result.iloc[: [0,1]]

print(short_sorted)







